from reportlab.lib.pagesizes import A4
from flask import Flask, render_template, request, redirect, send_file, session, url_for
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect, text
from copy import copy
import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Table
import os
import re

app = Flask(__name__)

# ===============================
# DATABASE CONFIG
# ===============================
database_url = os.environ.get("DATABASE_URL")

if database_url:
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql+psycopg://", 1)
    elif database_url.startswith("postgresql://"):
        database_url = database_url.replace("postgresql://", "postgresql+psycopg://", 1)

    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
    print("Using PostgreSQL database")
else:
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///groups.db"
    print("Using SQLite database")

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# If SECRET_KEY is missing or blank, Flask sessions break and admin login returns 500.
secret_key = (os.environ.get("SECRET_KEY") or "").strip()
app.config["SECRET_KEY"] = secret_key or "group-project-dev-secret"
db = SQLAlchemy(app)

ADMIN_PASSWORD = "1353"
MAX_GROUPS_PER_SUBJECT = 26
DEFAULT_SUBJECT_KEY = "microcontroller-interfacing"

SUBJECTS = [
    {
        "key": "microcontroller-interfacing",
        "name": "Microcontroller & Interfacing",
        "faculty": "Samiksha Gawali Mam",
        "deadline": "March 3, 2026 23:59:59",
        "topics": [
            "Smart Irrigation System",
            "Temperature Monitoring System",
            "Smart Parking System",
            "Home Automation using Arduino",
            "Digital Voltmeter",
            "Smart Dustbin",
            "Obstacle Avoiding Robot",
            "Heart Rate Monitor",
            "Weather Monitoring System",
            "Smart Energy Meter",
            "Voice Controlled Robot",
            "Automatic Plant Watering",
            "Solar Tracking System",
        ],
    },
    {
        "key": "digital-electronics",
        "name": "Digital Electronics",
        "faculty": "Ms. Shital Adasare Mam",
        "deadline": "March 7, 2026 23:59:59",
        "topics": [
            "Digital Parking Lot Occupancy Counter ",
            "Digital Door Lock System Using Logic Gates ",
            "Electronic Voting Machine Using Digital Logic ",
            "Binary Calculator Using Combinational Circuits ",
            "Automatic Traffic Light Controller Using Flip-Flops ",
            "Water Level Indicator Using Logic Gates ",
            "Binary to Decimal Converter Circuit ",
            "Gray Code to Binary Code Converter ",
            "4-bit Binary Adder Using Full Adders ",
            "Digital Counter Using Flip-Flops ",
            "Electronic Dice Using Digital Circuits ",
            "Binary Comparator Using Logic Gates ",
            "Burglar Alarm System Using Logic Circuits ",
            "Automatic Street Light Controller Using Digital Logic ",
            
        ],
    },
]
SUBJECTS_BY_KEY = {subject["key"]: subject for subject in SUBJECTS}


# ===============================
# MODEL
# ===============================
class Group(db.Model):
    __tablename__ = "groups"

    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(100), nullable=False, default=DEFAULT_SUBJECT_KEY, index=True)
    topic = db.Column(db.String(200))

    m1_name = db.Column(db.String(100))
    m1_prn = db.Column(db.String(100))
    m2_name = db.Column(db.String(100))
    m2_prn = db.Column(db.String(100))
    m3_name = db.Column(db.String(100))
    m3_prn = db.Column(db.String(100))
    m4_name = db.Column(db.String(100))
    m4_prn = db.Column(db.String(100))


class SubjectAccess(db.Model):
    __tablename__ = "subject_access"

    subject = db.Column(db.String(100), primary_key=True)
    is_open = db.Column(db.Boolean, nullable=False, default=True)


def ensure_subject_column():
    columns = [col["name"] for col in inspect(db.engine).get_columns("groups")]
    with db.engine.begin() as connection:
        if "subject" not in columns:
            connection.execute(text("ALTER TABLE groups ADD COLUMN subject VARCHAR(100)"))
        connection.execute(
            text(
                "UPDATE groups SET subject = :default_subject "
                "WHERE subject IS NULL OR subject = ''"
            ),
            {"default_subject": DEFAULT_SUBJECT_KEY},
        )


def migrate_sqlite_drop_topic_unique(connection):
    connection.execute(
        text(
            """
            CREATE TABLE groups_new (
                id INTEGER PRIMARY KEY,
                subject VARCHAR(100) NOT NULL,
                topic VARCHAR(200),
                m1_name VARCHAR(100),
                m1_prn VARCHAR(100),
                m2_name VARCHAR(100),
                m2_prn VARCHAR(100),
                m3_name VARCHAR(100),
                m3_prn VARCHAR(100),
                m4_name VARCHAR(100),
                m4_prn VARCHAR(100)
            )
            """
        )
    )
    connection.execute(
        text(
            """
            INSERT INTO groups_new (
                id, subject, topic, m1_name, m1_prn, m2_name, m2_prn, m3_name, m3_prn, m4_name, m4_prn
            )
            SELECT
                id,
                COALESCE(NULLIF(subject, ''), :default_subject),
                topic, m1_name, m1_prn, m2_name, m2_prn, m3_name, m3_prn, m4_name, m4_prn
            FROM groups
            """
        ),
        {"default_subject": DEFAULT_SUBJECT_KEY},
    )
    connection.execute(text("DROP TABLE groups"))
    connection.execute(text("ALTER TABLE groups_new RENAME TO groups"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_groups_subject ON groups (subject)"))


def ensure_topic_is_not_unique():
    unique_constraints = inspect(db.engine).get_unique_constraints("groups")
    topic_constraints = [
        constraint for constraint in unique_constraints if constraint.get("column_names") == ["topic"]
    ]
    if not topic_constraints:
        return

    with db.engine.begin() as connection:
        dialect = db.engine.dialect.name

        if dialect == "sqlite":
            migrate_sqlite_drop_topic_unique(connection)
            return

        for constraint in topic_constraints:
            constraint_name = constraint.get("name")
            if not constraint_name:
                continue
            if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", constraint_name):
                continue
            connection.execute(text(f"ALTER TABLE groups DROP CONSTRAINT {constraint_name}"))


def ensure_subject_access_rows():
    existing_subjects = {row.subject for row in SubjectAccess.query.all()}
    missing_subjects = [
        subject["key"] for subject in SUBJECTS if subject["key"] not in existing_subjects
    ]
    if not missing_subjects:
        return

    for subject_key in missing_subjects:
        db.session.add(SubjectAccess(subject=subject_key, is_open=True))
    db.session.commit()


with app.app_context():
    db.create_all()
    ensure_subject_column()
    ensure_topic_is_not_unique()
    ensure_subject_access_rows()


# ===============================
# HELPER FUNCTIONS
# ===============================
def clean_text(text):
    return re.sub(r"\s+", "", (text or "").lower())


def clean_words(text):
    text = (text or "").lower()
    text = re.sub(r"[^a-z\s]", "", text)
    return set(text.split())


def topics_similar(topic1, topic2):
    words1 = clean_words(topic1)
    words2 = clean_words(topic2)

    if not words1 or not words2:
        return False

    common = words1.intersection(words2)
    similarity_ratio = len(common) / min(len(words1), len(words2))
    return similarity_ratio >= 0.7


def normalize_subject_key(subject_key):
    key = (subject_key or "").strip().lower()
    if key in SUBJECTS_BY_KEY:
        return key
    return DEFAULT_SUBJECT_KEY


def get_selected_subject_key():
    raw_subject = request.values.get("subject")
    return normalize_subject_key(raw_subject)


def get_groups_for_subject(subject_key):
    return Group.query.filter_by(subject=subject_key).order_by(Group.id.asc()).all()


def get_subject_access_map():
    subject_access_map = {subject["key"]: True for subject in SUBJECTS}
    for row in SubjectAccess.query.all():
        subject_access_map[row.subject] = bool(row.is_open)
    return subject_access_map


def admin_required_redirect():
    if not session.get("is_admin"):
        return redirect(url_for("admin"))
    return None


# ===============================
# STUDENT PAGE
# ===============================
@app.route("/", methods=["GET", "POST"])
def index():
    popup = None
    message = None

    selected_subject_key = get_selected_subject_key()
    selected_subject = SUBJECTS_BY_KEY[selected_subject_key]
    all_groups_for_subject = get_groups_for_subject(selected_subject_key)
    subject_access_map = get_subject_access_map()
    selected_subject_open = subject_access_map.get(selected_subject_key, True)
    existing_groups = all_groups_for_subject if selected_subject_open else []
    all_topics = selected_subject["topics"]
    submitted_topics = {clean_text(g.topic) for g in existing_groups if g.topic}

    def render_index():
        return render_template(
            "index.html",
            groups=existing_groups,
            popup=popup,
            message=message,
            all_topics=all_topics,
            submitted_topics=submitted_topics,
            subjects=SUBJECTS,
            selected_subject=selected_subject,
            selected_subject_key=selected_subject_key,
            max_groups_per_subject=MAX_GROUPS_PER_SUBJECT,
            selected_subject_open=selected_subject_open,
            subject_access_map=subject_access_map,
        )

    if request.method == "POST":
        if not selected_subject_open:
            popup = "subject_closed"
            message = "Closed: Data is not visible because form is closed."
            return render_index()

        if Group.query.filter_by(subject=selected_subject_key).count() >= MAX_GROUPS_PER_SUBJECT:
            popup = "max_groups"
            message = f"Maximum {MAX_GROUPS_PER_SUBJECT} groups allowed for this subject."
            return render_index()

        topic = request.form.get("topic", "").strip()
        if not topic:
            popup = "invalid_topic"
            message = "Topic is required."
            return render_index()

        # ===============================
        # CHECK DUPLICATE TOPIC
        # ===============================
        for g in existing_groups:
            if topics_similar(topic, g.topic):
                popup = "duplicate_topic"
                message = f"Topic already selected by Group #{g.id} in this subject."
                return render_index()

        # ===============================
        # COLLECT MEMBERS (1-4)
        # ===============================
        members = []

        for i in range(1, 5):
            name = request.form.get(f"m{i}_name", "").strip()
            prn = request.form.get(f"m{i}_prn", "").strip()

            if name and prn:
                members.append((name, prn))

        if len(members) == 0:
            popup = "invalid_group"
            message = "At least 1 member is required."
            return render_index()

        # ===============================
        # PRN VALIDATION (12 digits)
        # ===============================
        for _name, prn in members:
            if not prn.isdigit() or len(prn) != 12:
                popup = "invalid_prn"
                message = f"PRN {prn} must be exactly 12 digits."
                return render_index()

        # ===============================
        # CHECK DUPLICATE MEMBERS
        # ===============================
        for g in existing_groups:
            existing_names = [g.m1_name, g.m2_name, g.m3_name, g.m4_name]
            existing_prns = [g.m1_prn, g.m2_prn, g.m3_prn, g.m4_prn]

            for name, prn in members:
                if clean_text(prn) in [clean_text(p) for p in existing_prns if p]:
                    popup = "duplicate_user"
                    message = f"PRN {prn} already in Group #{g.id} for this subject."
                    return render_index()

                if clean_text(name) in [clean_text(n) for n in existing_names if n]:
                    popup = "duplicate_user"
                    message = f"{name} already in Group #{g.id} for this subject."
                    return render_index()

        # ===============================
        # SAVE GROUP
        # ===============================
        new_group = Group(topic=topic, subject=selected_subject_key)

        if len(members) >= 1:
            new_group.m1_name, new_group.m1_prn = members[0]
        if len(members) >= 2:
            new_group.m2_name, new_group.m2_prn = members[1]
        if len(members) >= 3:
            new_group.m3_name, new_group.m3_prn = members[2]
        if len(members) >= 4:
            new_group.m4_name, new_group.m4_prn = members[3]

        db.session.add(new_group)
        db.session.commit()

        return redirect(url_for("index", subject=selected_subject_key))

    return render_index()


# ===============================
# ADMIN LOGIN + PANEL
# ===============================
@app.route("/admin", methods=["GET", "POST"])
def admin():
    if request.method == "POST" and "password" in request.form:
        if request.form["password"] == ADMIN_PASSWORD:
            session["is_admin"] = True
            return redirect(url_for("admin", subject=get_selected_subject_key()))
        return render_template("admin_login.html", error="Wrong Password!")

    if not session.get("is_admin"):
        return render_template("admin_login.html")

    if request.method == "POST" and request.form.get("action") == "set_subject_access":
        subject_key = normalize_subject_key(request.form.get("subject"))
        is_open = request.form.get("is_open") == "1"

        subject_access = db.session.get(SubjectAccess, subject_key)
        if subject_access is None:
            subject_access = SubjectAccess(subject=subject_key, is_open=is_open)
            db.session.add(subject_access)
        else:
            subject_access.is_open = is_open
        db.session.commit()

        return redirect(url_for("admin", subject=subject_key))

    selected_subject_key = get_selected_subject_key()
    groups = get_groups_for_subject(selected_subject_key)
    subject_access_map = get_subject_access_map()
    subject_is_open = subject_access_map.get(selected_subject_key, True)

    return render_template(
        "admin.html",
        groups=groups,
        subjects=SUBJECTS,
        selected_subject_key=selected_subject_key,
        selected_subject=SUBJECTS_BY_KEY[selected_subject_key],
        subject_access_map=subject_access_map,
        subject_is_open=subject_is_open,
    )


@app.route("/admin/logout", methods=["GET", "POST"])
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin"))


# ===============================
# ADMIN CRUD
# ===============================
@app.route("/admin/edit/<int:group_id>", methods=["GET", "POST"])
def edit_group(group_id):
    admin_redirect = admin_required_redirect()
    if admin_redirect:
        return admin_redirect

    group = Group.query.get_or_404(group_id)

    if request.method == "POST":
        selected_subject_key = normalize_subject_key(request.form.get("subject") or group.subject)
        topic = request.form.get("topic", "").strip()

        if topic:
            group.topic = topic
        group.subject = selected_subject_key

        for i in range(1, 5):
            name = request.form.get(f"m{i}_name", "").strip()
            prn = request.form.get(f"m{i}_prn", "").strip()
            setattr(group, f"m{i}_name", name or None)
            setattr(group, f"m{i}_prn", prn or None)

        db.session.commit()
        return redirect(url_for("admin", subject=selected_subject_key))

    selected_subject_key = normalize_subject_key(request.args.get("subject") or group.subject)
    return render_template(
        "edit_group.html",
        group=group,
        subjects=SUBJECTS,
        selected_subject_key=selected_subject_key,
    )


@app.route("/admin/delete/<int:group_id>", methods=["POST"])
def delete_group(group_id):
    admin_redirect = admin_required_redirect()
    if admin_redirect:
        return admin_redirect

    group = Group.query.get_or_404(group_id)
    selected_subject_key = normalize_subject_key(request.form.get("subject") or group.subject)
    db.session.delete(group)
    db.session.commit()
    return redirect(url_for("admin", subject=selected_subject_key))


# ===============================
# DOWNLOAD EXCEL (CIA FORMAT STYLE)
# ===============================
@app.route('/download_excel')
def download_excel():
    admin_redirect = admin_required_redirect()
    if admin_redirect:
        return admin_redirect

    from openpyxl import Workbook, load_workbook

    selected_subject_key = get_selected_subject_key()
    selected_subject = SUBJECTS_BY_KEY[selected_subject_key]
    groups = get_groups_for_subject(selected_subject_key)

    template_candidates = [
        os.environ.get("EXCEL_TEMPLATE_PATH"),
        os.path.join(app.root_path, "Microcontroller Interface CIA 3 Updated (1).xlsx"),
        r"C:\Users\patil\OneDrive\Documents\Microcontroller Interface CIA 3 Updated (1).xlsx",
    ]
    template_path = next((path for path in template_candidates if path and os.path.exists(path)), None)

    def clone_row_style(sheet, src_row, dst_row, max_col=4):
        src_dim = sheet.row_dimensions[src_row]
        if src_dim.height is not None:
            sheet.row_dimensions[dst_row].height = src_dim.height
        for col in range(1, max_col + 1):
            src = sheet.cell(src_row, col)
            dst = sheet.cell(dst_row, col)
            if src.has_style:
                dst._style = copy(src._style)

    def next_serial_number(sheet):
        serials = []
        for r in range(5, sheet.max_row + 1):
            value = sheet.cell(r, 1).value
            if value is None or value == "":
                continue
            try:
                serials.append(int(str(value).strip()))
            except ValueError:
                continue
        return (max(serials) + 1) if serials else 1

    def append_groups_like_template(sheet, start_row, groups_to_write, start_serial):
        row = start_row
        sr_no = start_serial
        style_rows = [5, 6, 7, 8] if sheet.max_row >= 8 else []

        for g in groups_to_write:
            members = []
            for name, prn in [
                (g.m1_name, g.m1_prn),
                (g.m2_name, g.m2_prn),
                (g.m3_name, g.m3_prn),
                (g.m4_name, g.m4_prn),
            ]:
                if name or prn:
                    members.append((name or "", prn or ""))

            if not members:
                members = [("", "")]

            for idx in range(len(members)):
                if style_rows:
                    clone_row_style(sheet, style_rows[min(idx, len(style_rows) - 1)], row + idx)

            if len(members) > 1:
                sheet.merge_cells(start_row=row, start_column=1, end_row=row + len(members) - 1, end_column=1)
                sheet.merge_cells(start_row=row, start_column=4, end_row=row + len(members) - 1, end_column=4)

            sheet.cell(row=row, column=1).value = sr_no
            sheet.cell(row=row, column=4).value = g.topic or ""

            for idx, (name, prn) in enumerate(members):
                current_row = row + idx
                sheet.cell(row=current_row, column=2).value = str(prn) if prn else ""
                sheet.cell(row=current_row, column=3).value = name

            row += len(members)
            sr_no += 1

    if template_path:
        wb = load_workbook(template_path)
        ws = wb.active

        if ws["A3"].value:
            ws["A3"] = f"CIA 3 {selected_subject['name']} list"

        start_row = ws.max_row + 1
        append_groups_like_template(
            ws,
            start_row,
            groups,
            start_serial=next_serial_number(ws),
        )
    else:
        from openpyxl.styles import Alignment, Font, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Mini Project List"

        ws.merge_cells('A1:D1')
        ws['A1'] = "Sandip University, Nashik (MS), India"
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A1'].font = Font(size=14, bold=True)

        ws.merge_cells('A2:D2')
        ws['A2'] = "Program: B.Tech CSE | Sem IV | Div A"
        ws['A2'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A3:D3')
        ws['A3'] = f"Subject: {selected_subject['name']}"
        ws['A3'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A4:D4')
        ws['A4'] = "Mini Project List"
        ws['A4'].alignment = Alignment(horizontal="center")
        ws['A4'].font = Font(size=13, bold=True)

        ws['A6'] = "Sr.No"
        ws['B6'] = "PRN No"
        ws['C6'] = "Project Group Members"
        ws['D6'] = "Project Title"

        for col in range(1, 5):
            ws.cell(row=6, column=col).font = Font(bold=True)

        row = 7
        sr = 1
        for g in groups:
            members = []
            for name, prn in [
                (g.m1_name, g.m1_prn),
                (g.m2_name, g.m2_prn),
                (g.m3_name, g.m3_prn),
                (g.m4_name, g.m4_prn),
            ]:
                if name or prn:
                    members.append((name or "", prn or ""))

            if not members:
                members = [("", "")]

            ws.cell(row=row, column=1).value = sr
            ws.cell(row=row, column=4).value = g.topic or ""

            for idx, (name, prn) in enumerate(members):
                current_row = row + idx
                ws.cell(row=current_row, column=2).value = str(prn) if prn else ""
                ws.cell(row=current_row, column=3).value = name

            row += len(members)
            sr += 1

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30

        thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for r in range(6, row):
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = thin

    safe_subject = selected_subject_key.replace("-", "_")
    file_path = f"groups_{safe_subject}.xlsx"
    wb.save(file_path)
    return send_file(
        file_path,
        as_attachment=True,
        download_name=f"{selected_subject['name']} Groups.xlsx",
    )


# ===============================
# DOWNLOAD PDF
# ===============================
@app.route("/download_pdf")
def download_pdf():
    admin_redirect = admin_required_redirect()
    if admin_redirect:
        return admin_redirect

    selected_subject_key = get_selected_subject_key()
    selected_subject = SUBJECTS_BY_KEY[selected_subject_key]
    groups = get_groups_for_subject(selected_subject_key)

    safe_subject = selected_subject_key.replace("-", "_")
    file_path = f"groups_{safe_subject}.pdf"
    doc = SimpleDocTemplate(file_path, pagesize=A4)

    from reportlab.platypus import Paragraph, Spacer, TableStyle
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib import colors

    elements = []
    styles = getSampleStyleSheet()

    elements.append(Spacer(1, 40))

    university_style = ParagraphStyle(
        "UniversityStyle",
        parent=styles["Title"],
        alignment=1,
        fontSize=18,
        spaceAfter=10,
    )

    normal_center = ParagraphStyle(
        "NormalCenter",
        parent=styles["Normal"],
        alignment=1,
        fontSize=12,
    )

    main_heading = ParagraphStyle(
        "MainHeading",
        parent=styles["Heading1"],
        alignment=1,
        fontSize=20,
        spaceBefore=15,
        spaceAfter=20,
    )

    elements.append(Paragraph("<b>Sandip University, Nashik (MS), India</b>", university_style))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("Program: B.Tech CSE", normal_center))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph("Sem IV | Div A", normal_center))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(f"Subject: {selected_subject['name']}", normal_center))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(f"Faculty: {selected_subject['faculty']}", normal_center))
    elements.append(Spacer(1, 25))
    elements.append(Paragraph("<b>Mini Project List</b>", main_heading))
    elements.append(Spacer(1, 15))

    data = [["Sr.no", "PRN No", "Project group members", "Project title"]]

    for sr, g in enumerate(groups, start=1):
        prns = []
        names = []

        if g.m1_prn:
            prns.append(f"1) {g.m1_prn}")
        if g.m2_prn:
            prns.append(f"2) {g.m2_prn}")
        if g.m3_prn:
            prns.append(f"3) {g.m3_prn}")
        if g.m4_prn:
            prns.append(f"4) {g.m4_prn}")

        if g.m1_name:
            names.append(g.m1_name)
        if g.m2_name:
            names.append(g.m2_name)
        if g.m3_name:
            names.append(g.m3_name)
        if g.m4_name:
            names.append(g.m4_name)

        data.append(
            [
                f"{sr})",
                "\n".join(prns),
                "\n".join(names),
                g.topic or "",
            ]
        )

    table = Table(data, colWidths=[40, 120, 160, 150])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)

    return send_file(
        file_path,
        as_attachment=True,
        download_name=f"{selected_subject['name']} Groups.pdf",
    )


if __name__ == "__main__":
    app.run(debug=True)
